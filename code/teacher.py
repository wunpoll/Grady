from PyQt6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                            QListWidget, QLabel, QPushButton, QScrollArea, QFileDialog)
import sqlite3
from openpyxl import Workbook
from datetime import datetime
from PyQt6.QtGui import QIcon, QColor, QFont
from PyQt6.QtCore import QSize
import os
import sys

def resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, 'media/', relative_path)

class TeacherWindow(QMainWindow):
    def __init__(self, teacher_id):
        super().__init__()
        self.teacher_id = teacher_id
        self.setWindowTitle("Панель учителя")
        self.setGeometry(100, 100, 800, 600)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        content_layout = QHBoxLayout()
        main_layout.addLayout(content_layout)

        students_header_container = QWidget()
        students_header_container.setFixedWidth(300)  
        students_header_layout = QHBoxLayout(students_header_container)
        students_header_layout.setContentsMargins(5, 5, 5, 5)

        students_title = QLabel("Список учеников")
        students_title.setFont(QFont('', 12, QFont.Weight.Bold))
        students_header_layout.addWidget(students_title)

        refresh_button = QPushButton()
        refresh_button.setIcon(QIcon.fromTheme("view-refresh"))
        refresh_button.setFixedSize(24, 24)
        refresh_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                border: none;
                border-radius: 4px;
                padding: 4px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        refresh_button.setToolTip("Обновить список учеников")
        refresh_button.clicked.connect(self.load_students)
        students_header_layout.addWidget(refresh_button)

        left_container = QVBoxLayout()
        left_container.addWidget(students_header_container)

        self.students_list = QListWidget()
        self.students_list.setMaximumWidth(300)
        self.load_students()
        self.students_list.itemClicked.connect(self.show_student_info)
        left_container.addWidget(self.students_list)

        content_layout.addLayout(left_container)

        self.info_widget = QWidget()
        self.info_layout = QVBoxLayout(self.info_widget)
        scroll = QScrollArea()
        scroll.setWidget(self.info_widget)
        scroll.setWidgetResizable(True)
        content_layout.addWidget(scroll)

        self.export_button = QPushButton("Сформировать отчет Excel")
        self.export_button.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                padding: 10px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        self.export_button.clicked.connect(self.generate_excel_report)
        main_layout.addWidget(self.export_button)

    def generate_excel_report(self):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчет по успеваемости"
            headers = ["Фамилия", "Имя", "Отчество", "Образование матери", 
                      "Образование отца", "Свободное время (ч)", 
                      "Доп. занятия", "Участие в олимпиадах", "KPI", "Оценка"]
            ws.append(headers)

            db_path = resource_path("Grady.db")
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            cursor.execute("""
                SELECT group_id, Groups.name 
                FROM Teacher 
                JOIN Groups ON Teacher.group_id = Groups.id 
                WHERE user_id = ?
            """, (self.teacher_id,))
            group_data = cursor.fetchone()
            group_id, group_name = group_data

            cursor.execute("""
                SELECT 
                    Student.last_name,
                    Student.first_name,
                    Student.middle_name,
                    Factors.mother_education,
                    Factors.father_education,
                    Factors.free_time_hours,
                    Factors.additional_activities,
                    Factors.olympiads_part,
                    Student.user_id,
                    Grades.predicted_grade
                FROM Student
                LEFT JOIN Factors ON Student.user_id = Factors.student_id
                LEFT JOIN Grades ON Student.user_id = Grades.student_id
                WHERE Student.group_id = ?
            """, (group_id,))
            
            students_data = cursor.fetchall()
            print(students_data)
            for student in students_data:
                kpi = student[9] if student[9] is not None else 0
                grade = self.convert_kpi_to_grade(kpi)
                row_data = list(student[:8]) + [kpi, grade]
                ws.append(row_data)

            for column in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + column)].width = 15
            ws.insert_rows(1, 3)
            ws['A1'] = f"Группа: {group_name}"
            ws['A2'] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить отчет",
                f"Отчет_группа_{group_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if file_path:
                wb.save(file_path)
                self.show_success_message("Отчет успешно сохранен!")

            conn.close()

        except Exception as e:
            self.show_error_message(f"Ошибка при создании отчета: {str(e)}")

    def show_success_message(self, message):
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.information(self, "Успех", message)

    def show_error_message(self, message):
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.critical(self, "Ошибка", message)

    def load_students(self):
        
        db_path = resource_path("Grady.db")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT group_id FROM Teacher WHERE user_id = ?
        """, (self.teacher_id,))
        group_id = cursor.fetchone()[0]
        
        cursor.execute("""
            SELECT Student.user_id, Student.last_name, Student.first_name, Student.middle_name 
            FROM Student 
            WHERE Student.group_id = ?
        """, (group_id,))
        
        students = cursor.fetchall()
        self.students_list.clear()
        for student in students:
            self.students_list.addItem(f"{student[1]} {student[2]} {student[3]}")
        
        conn.close()

    def show_student_info(self, item):
        while self.info_layout.count():
            child = self.info_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()


        db_path = resource_path("Grady.db")
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        student_name = item.text().split()
        cursor.execute("""
            SELECT user_id FROM Student 
            WHERE last_name = ? AND first_name = ? AND middle_name = ?
        """, (student_name[0], student_name[1], student_name[2]))
        student_id = cursor.fetchone()[0]

        cursor.execute("""
            SELECT mother_education, father_education, free_time_hours, 
                   additional_activities, olympiads_part 
            FROM Factors 
            WHERE student_id = ?
        """, (student_id,))
        factors = cursor.fetchone()

        if factors:
            self.info_layout.addWidget(QLabel(f"Образование матери: {factors[0]}"))
            self.info_layout.addWidget(QLabel(f"Образование отца: {factors[1]}"))
            self.info_layout.addWidget(QLabel(f"Свободное время (часов): {factors[2]}"))
            self.info_layout.addWidget(QLabel(f"Доп. занятия: {factors[3]}"))
            self.info_layout.addWidget(QLabel(f"Участие в олимпиадах: {factors[4]}"))

            cursor.execute("""
                SELECT predicted_grade 
                FROM Grades 
                WHERE student_id = ?
            """, (student_id,))
            grade_data = cursor.fetchone()
            
            if grade_data:
                kpi = grade_data[0]
                grade = self.convert_kpi_to_grade(kpi)
                
                kpi_label = QLabel(f"KPI: {kpi}")
                grade_label = QLabel(f"Предполагаемая оценка: {grade}")
                
                kpi_color = self.get_color_for_kpi(kpi)
                grade_color = self.get_color_for_grade(grade)
            
            kpi_label.setStyleSheet(f"QLabel {{ color: {kpi_color}; font-weight: bold; font-size: 14px; }}")
            grade_label.setStyleSheet(f"QLabel {{ color: {grade_color}; font-weight: bold; font-size: 14px; }}")
            
            self.info_layout.addWidget(QLabel(""))
            self.info_layout.addWidget(kpi_label)
            self.info_layout.addWidget(grade_label)

        conn.close()

    def convert_kpi_to_grade(self, kpi):
        if kpi >= 80: return 5
        elif kpi >= 60: return 4
        elif kpi >= 40: return 3
        else: return 2

    def get_color_for_kpi(self, kpi):
        if kpi >= 80:
            return "#2ecc71" 
        elif kpi >= 60:
            return "#f1c40f" 
        elif kpi >= 40:
            return "#e67e22" 
        else:
            return "#e74c3c"

    def get_color_for_grade(self, grade):
        if grade == 5:
            return "#2ecc71"  
        elif grade == 4:
            return "#f1c40f"   
        elif grade == 3:
            return "#e67e22"  
        else:
            return "#e74c3c"  

